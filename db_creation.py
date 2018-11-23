# -*- coding: utf-8 -*-
"""
_author_    = "Ricardo Lucas"
_copyright_ = "Copyright 2018, ESTIG - IPBeja"
_version_   = "13:05 02/02/2018"
"""

from app_creation import db
import os.path
from openpyxl import load_workbook

# Path to data.xlsx
path_data = 'C:/Users/Ricardo/Desktop/LP/PYTHON/Lucas15297/lp15297/src/data/data.xlsx'

# Loading WorkBook from the path_data
wb_data = load_workbook(path_data)

print('Creating uniMadeira.db:')


class Universidades(db.Model):
    '''
    Criação da Table Universidades

    Args:
        :arg (obj): Object Model

    Atributes:
        nome (str): Nome da Universidade
        historia (str)= Historia da Universidade
        ano_origem (int) = Ano em que a Universidade foi inaugurada
        reitor (str)= Nome do Reitor
        localidade (str)= Localidade emq ue está situada a Universidade
        morada (str) = Morada da Universidade
        telefone (BigInt) = Telefone da Universidade
        fax (BigInt) = Fax da Universidade
        email_uni (str) = Email da Universidade
        email_reitoria (str) = Email do Reitor da Universidade
        horario_atendimento (str) = Horario de Atendimento
        requisitos_matricula (str) = requisitos Matrícula

    '''
    __tablename__ = 'universidades'
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String)
    historia = db.Column(db.String)
    ano_origem = db.Column(db.Integer)
    reitor = db.Column(db.String)
    localidade = db.Column(db.String)
    morada = db.Column(db.String)
    telefone = db.Column(db.BigInteger)
    fax = db.Column(db.BigInteger)
    email_uni = db.Column(db.String)
    email_reitoria = db.Column(db.String)
    horario_atendimento = db.Column(db.String)
    requisitos_matricula = db.Column(db.String)

    def __init__(self, nome, historia, ano_origem, reitor, localidade, morada, telefone, fax, email_uni, email_reitoria,
                 horario_atendimento, requisitos_matricula):
        self.nome = nome
        self.historia = historia
        self.ano_origem = ano_origem
        self.reitor = reitor
        self.localidade = localidade
        self.morada = morada
        self.telefone = telefone
        self.fax = fax
        self.email_uni = email_uni
        self.email_reitoria = email_reitoria
        self.horario_atendimento = horario_atendimento
        self.requisitos_matricula = requisitos_matricula

class Escolas(db.Model):
    '''
        Criação da Table Escolas

        Args:
            :arg (obj): Object Model

        Atributes:
            nome_escola (str): Nome da Escola
            descricao_escola (str) = Descriçao da Escola
            nome_presidente (str) = Nome do Presidente
            nome_vice_presidente (str)= Nome do vice Presidente
            localidade (str)= Localidade emq ue está situada a Universidade
            conselho_cientifico (str) = Conselho Cientifico
            universidade_id (BigInt) = Id da Universidade
            fax (BigInt) = Fax da Universidade
            email_uni (str) = Email da Universidade
            email_reitoria (str) = Email do Reitor da Universidade
            horario_atendimento (str) = Horario de Atendimento
            requisitos_matricula (str) = requisitos Matrícula

        '''
    __tablename__ = 'escolas'
    id = db.Column(db.Integer, primary_key=True)
    nome_escola = db.Column(db.String)
    descricao_escola = db.Column(db.String)
    nome_presidente = db.Column(db.String)
    nome_vice_presidente = db.Column(db.String)
    conselho_cientifico = db.Column(db.String)
    universidade_id = db.Column(db.Integer, db.ForeignKey('universidades.id'))
    ligacao = db.relationship("Universidades", backref=db.backref('escolas', order_by=id))

    def __init__(self, nome_escola, descricao_escola, nome_presidente, nome_vice_presidente, conselho_cientifico,
                 universidade):
        self.nome_escola = nome_escola
        self.descricao_escola = descricao_escola
        self.nome_presidente = nome_presidente
        self.nome_vice_presidente = nome_vice_presidente
        self.conselho_cientifico = conselho_cientifico
        self.universidade_id = universidade

class Laboratorios(db.Model):
    '''
        Criação da Table Laboratorios

        Args:
            :arg (obj): Object Model

        Atributes:
            nome_laboratorio (str): Nome  Laboratorio
            quantidade (int) = Quantidade de Laboratorios
            universidade_id (int) = Id da Universidade
            nome_vice_presidente (str)= Nome do vice Presidente
            localidade (str)= Localidade emq ue está situada a Universidade
            conselho_cientifico (str) = Conselho Cientifico
            universidade_id (BigInt) = Id da Universidade
            fax (BigInt) = Fax da Universidade
            email_uni (str) = Email da Universidade
            email_reitoria (str) = Email do Reitor da Universidade
            horario_atendimento (str) = Horario de Atendimento
            requisitos_matricula (str) = requisitos Matrícula

    '''
    __tablename__ = 'laboratorios'
    id = db.Column(db.Integer, primary_key=True)
    nome_laboratorio = db.Column(db.String)
    quantidade = db.Column(db.Integer)
    universidade_id = db.Column(db.Integer, db.ForeignKey('universidades.id'))
    ligacao = db.relationship("Universidades", backref=db.backref('laboratorios', order_by=id))

    def __init__(self, nome_laboratorio, quantidade, universidade):
        self.nome_laboratorio = nome_laboratorio
        self.quantidade = quantidade
        self.universidade_id = universidade

class Cursos(db.Model):
    '''
        Criação da Table Laboratorios

        Args:
            :arg (obj): Object Model

        Atributes:
            nome_curso (str): Nome Curso
            tipo_curso (str) = Tipo de Curso
            cod_escola (int) = Código Instituição
            nome_vice_presidente (str)= Nome do vice Presidente
            duracao_curso (str)= Duração do Curso
            nome_escola (str) = Nome Escola
            saida_descricao (String) = Saidas Descriçao
            nome_director (String) = Nome do Director do curso
            email_director (str) = Email do Director
            objectivos_curso (str) = Objectivos do Curso
            escola_id (Int) = Id da Escola
            requisitos_matricula (str) = requisitos Matrícula
    '''

    __tablename__ = 'cursos'
    id = db.Column(db.Integer, primary_key=True)
    nome_curso = db.Column(db.String)
    tipo_curso = db.Column(db.String)
    cod_escola = db.Column(db.Integer)
    duracao_curso = db.Column(db.String)
    nome_escola = db.Column(db.String)
    saidas_descricao = db.Column(db.String)
    nome_director = db.Column(db.String)
    email_director = db.Column(db.String)
    objectivos_curso = db.Column(db.String)
    escola_id = db.Column(db.Integer, db.ForeignKey('escolas.id'))
    ligacao = db.relationship("Escolas", backref=db.backref('cursos', order_by=id))

    def __init__(self, nome_curso, tipo_curso, cod_escola, duracao_curso, nome_escola, saidas_descricao, nome_director,
                 email_director, objectivos_curso, escola):
        self.nome_curso = nome_curso
        self.tipo_curso = tipo_curso
        self.cod_escola = cod_escola
        self.duracao_curso = duracao_curso
        self.nome_escola = nome_escola
        self.saidas_descricao = saidas_descricao
        self.nome_director = nome_director
        self.email_director = email_director
        self.objectivos_curso = objectivos_curso
        self.escola_id = escola

class unidadesCurriculares(db.Model):
    '''
    Criação da Table unidadesCurriculares

    Args:
        :arg (obj): Object Model

    Atributes:
        nome_disciplina (str)= Nome da disciplina
        area_cientifica (str)= Nome area Cientifica
        nivel (str)= Nivel
        ficha_descritiva (str)= Ficha descritiva
        disciplina_ativa (str)= Disciplina activa
        ano_curricular (int)= Ano Curricular
        semestre (str)= Semestre
        horas_teoricas_semanal (int)= Horas Teoricas Semanais
        horas_teorica_praticas_semanal (int)= Horas teorica praticas semanais
        horas_praticas_laboratorial (int)= Horas pratica laboratorial
        horas_trabalho_campo (int)= horas trabalho campo
        horas_seminario_semanal (int)=horas seminario semanal
        horas_estagio_semanal (int)= horas estagio semanal
        horas_orientacao_tutorial_semanal(int) = horas orientacao Tutorial semanal
        creditos_disciplina (int)= Creditos da Disciplina
        curso_id (int)= Id do Curso
    '''
    __tablename__ = 'unidades_curriculares'
    id = db.Column(db.Integer, primary_key=True)
    nome_disciplina = db.Column(db.String)
    area_cientifica = db.Column(db.String)
    nivel = db.Column(db.String)
    ficha_descritiva = db.Column(db.String)
    disciplina_ativa = db.Column(db.String)
    ano_curricular = db.Column(db.Integer)
    semestre = db.Column(db.String)
    horas_teoricas_semanal = db.Column(db.Integer)
    horas_teorica_praticas_semanal = db.Column(db.Integer)
    horas_praticas_laboratorial = db.Column(db.Integer)
    horas_trabalho_campo = db.Column(db.Integer)
    horas_seminario_semanal = db.Column(db.Integer)
    horas_estagio_semanal = db.Column(db.Integer)
    horas_orientacao_tutorial_semanal = db.Column(db.Integer)
    creditos_disciplina = db.Column(db.Integer)
    curso_id = db.Column(db.Integer, db.ForeignKey('cursos.id'))
    ligacao = db.relationship("Cursos", backref=db.backref('unidades_curriculares', order_by=id))

    def __init__(self, nome_disciplina, area_cientifica, nivel, ficha_descritiva, disciplina_ativa, ano_curricular,
                 semestre, horas_teoricas_semanal, horas_teorica_praticas_semanal, horas_praticas_laboratorial,
                 horas_trabalho_campo, horas_seminario_semanal, horas_estagio_semanal,
                 horas_orientacao_tutorial_semanal, creditos_disciplina, curso):
        self.nome_disciplina = nome_disciplina
        self.area_cientifica = area_cientifica
        self.nivel = nivel
        self.ficha_descritiva = ficha_descritiva
        self.disciplina_ativa = disciplina_ativa
        self.ano_curricular = ano_curricular
        self.semestre = semestre
        self.horas_teoricas_semanal = horas_teoricas_semanal
        self.horas_teorica_praticas_semanal = horas_teorica_praticas_semanal
        self.horas_praticas_laboratorial = horas_praticas_laboratorial
        self.horas_trabalho_campo = horas_trabalho_campo
        self.horas_seminario_semanal = horas_seminario_semanal
        self.horas_estagio_semanal = horas_estagio_semanal
        self.horas_orientacao_tutorial_semanal = horas_orientacao_tutorial_semanal
        self.creditos_disciplina = creditos_disciplina
        self.curso_id = curso

class Professores(db.Model):
    '''
        Criação da Table Professores

        Args:
            :arg (obj): Object Model

        Atributes:
            nome_professor (str)= Nome do Professor
            email_professor (str)= Email do Professor
            habilitacoes (str)= Habilitacoes do professor
            cadeiras_lecionadas (str)= Cadeiras Lecionadas pelo Professor
            responsabilidades (str)= Responsabilidades do Professor
            actividade_profissional (str)= Actividade do Professor
            media_salario (int)= Media de salário
            area_pesquisa (str)= Area de Pesquisa
            licenciatura (str)= Licenciatura
            mestrado (str)= Mestrado
            doutorameno (str)= Doutoramento
        '''
    __tablename__ = 'professores'
    id = db.Column(db.Integer, primary_key=True)
    nome_professor = db.Column(db.String)
    email_professor = db.Column(db.String)
    habilitacoes = db.Column(db.String)
    cadeiras_lecionadas = db.Column(db.Integer)
    responsabilidades = db.Column(db.String)
    actividade_profissional = db.Column(db.String)
    media_salario = db.Column(db.BigInteger)
    area_pesquisa = db.Column(db.String)
    licenciatura = db.Column(db.String)
    mestrado = db.Column(db.String)
    doutoramento = db.Column(db.String)

    def __init__(self, nome_professor, email_professor, habilitacoes, cadeiras_lecionadas, responsabilidades,
                 actividade_profissional, media_salario, area_pesquisa, licenciatura, mestrado, doutoramento):
        self.nome_professor = nome_professor
        self.email_professor = email_professor
        self.habilitacoes = habilitacoes
        self.cadeiras_lecionadas = cadeiras_lecionadas
        self.responsabilidades = responsabilidades
        self.actividade_profissional = actividade_profissional
        self.media_salario = media_salario
        self.area_pesquisa = area_pesquisa
        self.licenciatura = licenciatura
        self.mestrado = mestrado
        self.doutoramento = doutoramento

class unidadesCurricularesProfessores(db.Model):
    '''
    Criação da Table PunidadesCurricularesProfessores

        Args:
            :arg (obj): Object Model

        Atributes:
            professor_id (int)= Id do professor
            cadeira_id (int)= Id da Cadeira
        '''
    __tablename__ = 'professores_cadeiras'
    id = db.Column(db.Integer, primary_key=True)
    professor_id = db.Column(db.Integer, db.ForeignKey('professores.id'))
    ligacao = db.relationship("Professores", backref=db.backref('professores_cadeiras', order_by=id))
    cadeiras_id = db.Column(db.Integer, db.ForeignKey('unidades_curriculares.id'))
    ligacao = db.relationship("unidadesCurriculares", backref=db.backref('professores_cadeiras', order_by=id))

    def __init__(self, professor, cadeira):
        self.professor_id = professor
        self.cadeiras_id = cadeira

class acessoCursos(db.Model):
    '''
    Criação da Table acessoCursos

        Args:
            :arg (obj): Object Model

        Atributes:
            ano (int)= Ano
            fase (int)= Fase
            cod_curso (int)= Código de Curso
            curso_id (int)= Id do Curso
    '''
    __tablename__ = 'acesso_cursos'
    id = db.Column(db.Integer, primary_key=True)
    ano = db.Column(db.Integer)
    fase = db.Column(db.Integer)
    cod_curso = db.Column(db.Integer)
    curso_id = db.Column(db.Integer, db.ForeignKey('cursos.id'))
    ligacao = db.relationship("Cursos", backref=db.backref('acesso_cursos', order_by=id))

    def __init__(self, ano, fase, cod_curso, curso):
        self.ano = ano
        self.fase = fase
        self.cod_curso = cod_curso
        self.curso_id = curso

class opcaoCandidatura(db.Model):
    '''
        Criação da Table opcaoCandidatura

            Args:
                :arg (obj): Object Model

            Atributes:
                opcao (int)= Opçao
                candidatos (int)= Candidatos
                percentagem_candidatos (int)= Percentagem Candidatos
                colocados (int)= Colocados
                percentagem_colocados (int)= Percentagem de Colocados
                acesso_id (int)= Id de Acessos
        '''
    __tablename__ = 'opcao_candidatura'
    id = db.Column(db.Integer, primary_key=True)
    opcao = db.Column(db.Integer)
    candidatos = db.Column(db.Integer)
    percentagem_candidatos = db.Column(db.Integer)
    colocados = db.Column(db.Integer)
    percentagem_colocados = db.Column(db.Integer)
    acesso_id = db.Column(db.Integer, db.ForeignKey('acesso_cursos.id'))
    ligacao = db.relationship("acessoCursos", backref=db.backref('opcao_candidatura', order_by=id))

    def __init__(self, opcao, candidatos, percentagem_candidatos, colocados, percentagem_colocados, acesso):
        self.opcao = opcao
        self.candidatos = candidatos
        self.percentagem_candidatos = percentagem_candidatos
        self.colocados = colocados
        self.percentagem_colocados = percentagem_colocados
        self.acesso_id = acesso

class etapaColocacao(db.Model):
    '''
        Criação da Table opcaoColocação

            Args:
                :arg (obj): Object Model

            Atributes:
                etapa (str)= Etapa
                candidatos (int)= Candidatos
                percentagem_candidatos (int)= Percentagem Candidatos
                colocados (int)= Colocados
                percentagem_colocados (int)= Percentagem de Colocados
                nota (int) = Nota
                acesso_id (int)= Id de Acessos
            '''
    __tablename__ = 'etapa_colocacao'
    id = db.Column(db.Integer, primary_key=True)
    etapa = db.Column(db.String)
    candidatos = db.Column(db.Integer)
    percentagem_candidatos = db.Column(db.Integer)
    colocados = db.Column(db.Integer)
    percentagem_colocados = db.Column(db.Integer)
    nota = db.Column(db.Integer)
    acesso_id = db.Column(db.Integer, db.ForeignKey('acesso_cursos.id'))
    ligacao = db.relationship("acessoCursos", backref=db.backref('etapa_colocacao', order_by=id))

    def __init__(self, etapa, candidatos, percentagem_candidatos, colocados, percentagem_colocados, nota, acesso):
        self.etapa = etapa
        self.candidatos = candidatos
        self.percentagem_candidatos = percentagem_candidatos
        self.colocados = colocados
        self.percentagem_colocados = percentagem_colocados
        self.nota = nota
        self.acesso_id = acesso

class distritoCandidatura(db.Model):
    '''
        Criação da Table distritoCandidatura

            Args:
                :arg (obj): Object Model

            Atributes:
                distrito (str)= Distrito
                candidatos (int)= Candidatos
                percentagem_candidatos (int)= Percentagem Candidatos
                colocados (int)= Colocados
                percentagem_colocados (int)= Percentagem de Colocados
                acesso_id (int)= Id de Acessos
    '''
    __tablename__ = 'distrito_candidatura'
    id = db.Column(db.Integer, primary_key=True)
    distrito = db.Column(db.String)
    candidatos = db.Column(db.Integer)
    percentagem_candidatos = db.Column(db.Integer)
    colocados = db.Column(db.Integer)
    percentagem_colocados = db.Column(db.Integer)
    acesso_id = db.Column(db.Integer, db.ForeignKey('acesso_cursos.id'))
    ligacao = db.relationship("acessoCursos", backref=db.backref('distrito_candidatura', order_by=id))

    def __init__(self, distrito, candidatos, percentagem_candidatos, colocados, percentagem_colocados, acesso):
        self.distrito = distrito
        self.candidatos = candidatos
        self.percentagem_candidatos = percentagem_candidatos
        self.colocados = colocados
        self.percentagem_colocados = percentagem_colocados
        self.acesso_id = acesso

class sexoCandidatura(db.Model):
    '''
        Criação da Table sexoCandidatura

            Args:
                :arg (obj): Object Model

            Atributes:
                sexo (str)= Género
                candidatos (int)= Candidatos
                percentagem_candidatos (int)= Percentagem Candidatos
                colocados (int)= Colocados
                percentagem_colocados (int)= Percentagem de Colocados
                nota (int) = Nota
                acesso_id (int)= Id de Acessos
    '''
    __tablename__ = 'sexo_candidatura'
    id = db.Column(db.Integer, primary_key=True)
    sexo = db.Column(db.String)
    candidatos = db.Column(db.Integer)
    percentagem_candidatos = db.Column(db.Integer)
    colocados = db.Column(db.Integer)
    percentagem_colocados = db.Column(db.Integer)
    acesso_id = db.Column(db.Integer, db.ForeignKey('acesso_cursos.id'))
    ligacao = db.relationship("acessoCursos", backref=db.backref('sexo_candidatura', order_by=id))

    def __init__(self, sexo, candidatos, percentagem_candidatos, colocados, percentagem_colocados, acesso):
        self.sexo = sexo
        self.candidatos = candidatos
        self.percentagem_candidatos = percentagem_candidatos
        self.colocados = colocados
        self.percentagem_colocados = percentagem_colocados
        self.acesso_id = acesso

class cursoCandidatos(db.Model):
    '''
        Criação da Table opcaoCandidatos

            Args:
                :arg (obj): Object Model

            Atributes:
                curso (str)= Curso
                candidatos (int)= Candidatos
                percentagem_candidatos (int)= Percentagem Candidatos
                colocados (int)= Colocados
                percentagem_colocados (int)= Percentagem de Colocados
                nota (int) = Nota
                acesso_id (int)= Id de Acessos
    '''
    __tablename__ = 'curso_candidatos'
    id = db.Column(db.Integer, primary_key=True)
    curso = db.Column(db.String)
    candidatos = db.Column(db.Integer)
    percentagem_candidatos = db.Column(db.Integer)
    colocados = db.Column(db.Integer)
    percentagem_colocados = db.Column(db.Integer)
    acesso_id = db.Column(db.Integer, db.ForeignKey('acesso_cursos.id'))
    ligacao = db.relationship("acessoCursos", backref=db.backref('curso_candidatos', order_by=id))

    def __init__(self, curso, candidatos, percentagem_candidatos, colocados, percentagem_colocados, acesso):
        self.curso = curso
        self.candidatos = candidatos
        self.percentagem_candidatos = percentagem_candidatos
        self.colocados = colocados
        self.percentagem_colocados = percentagem_colocados
        self.acesso_id = acesso

class mediaColocados(db.Model):
    '''
        Criação da Table mediaColocados

            Args:
                :arg (obj): Object Model

            Atributes:
                tipo_media (str)= Tipo Media
                media (int)= Média
                acesso_id (int)= Ids Asesso
                colocados (int)= Colocados
                percentagem_colocados (int)= Percentagem de Colocados
                nota (int) = Nota
                acesso_id (int)= Id de Acessos
                '''
    __tablename__ = 'media_colocados'
    id = db.Column(db.Integer, primary_key=True)
    tipo_media = db.Column(db.String)
    media = db.Column(db.Integer)
    acesso_id = db.Column(db.Integer, db.ForeignKey('acesso_cursos.id'))
    ligacao = db.relationship("acessoCursos", backref=db.backref('media_colocados', order_by=id))

    def __init__(self, tipo_media, media, acesso):
        self.tipo_media = tipo_media
        self.media = media
        self.acesso_id = acesso

class opcoesExcluidas(db.Model):
    '''
        Criação da Table opcaoExcluidas

            Args:
                :arg (obj): Object Model

            Atributes:
                tipo_exclusao (str)= Tipo de Exclusao
                quantidade (int)= Quantidade
                acesso_id (int)= Ids Asesso
                colocados (int)= Colocados
                percentagem_colocados (int)= Percentagem de Colocados
                nota (int) = Nota
                acesso_id (int)= Id de Acessos
    '''
    __tablename__ = 'opcoes_excluidas'
    id = db.Column(db.Integer, primary_key=True)
    tipo_exclusao = db.Column(db.String)
    quantidade = db.Column(db.Integer)
    acesso_id = db.Column(db.Integer, db.ForeignKey('acesso_cursos.id'))
    ligacao = db.relationship("acessoCursos", backref=db.backref('opcoes_excluidas', order_by=id))

    def __init__(self, tipo_exclusao, quantidade, acesso):
        self.tipo_exclusao = tipo_exclusao
        self.quantidade = quantidade
        self.acesso_id = acesso

class notasAlunosCandidatos(db.Model):
    '''
        Criação da Table notasAlunosCandidatos

            Args:
                :arg (obj): Object Model

            Atributes:
                nota (int)= Nota
                quantidade (int)= Quantidade
                acesso_id (int)= Id de Acessos
        '''
    __tablename__ = 'notas_alunos_candidatos'
    id = db.Column(db.Integer, primary_key=True)
    nota = db.Column(db.Integer)
    quantidade = db.Column(db.Integer)
    acesso_id = db.Column(db.Integer, db.ForeignKey('acesso_cursos.id'))
    ligacao = db.relationship("acessoCursos", backref=db.backref('notas_alunos_candidatos', order_by=id))

    def __init__(self, nota, quantidade, acesso):
        self.nota = nota
        self.quantidade = quantidade
        self.acesso_id = acesso

class notasAlunosColocados(db.Model):
    '''
        Criação da Table notasAlunosColocados

            Args:
                :arg (obj): Object Model

            Atributes:
                nota (int)= Nota
                quantidade (int)= Quantidade
                acesso_id (int)= Ids Asesso

    '''
    __tablename__ = 'notas_alunos_colocados'
    id = db.Column(db.Integer, primary_key=True)
    nota = db.Column(db.Integer)
    quantidade = db.Column(db.Integer)
    acesso_id = db.Column(db.Integer, db.ForeignKey('acesso_cursos.id'))
    ligacao = db.relationship("acessoCursos", backref=db.backref('notas_alunos_colocados', order_by=id))

    def __init__(self, nota, quantidade, acesso):
        self.nota = nota
        self.quantidade = quantidade
        self.acesso_id = acesso

#Create all tables in db
db.create_all()
print("Tables Created")


class insertUniversidade:
    '''
    Inserção de Dados na Tabela Universidades apartir do ficheiro informations
    '''

    sheet = wb_data.get_sheet_by_name('Universidade_data')
    row_count = sheet.max_row

    for row in range(1, row_count + 1):
        nome = sheet.cell(None, row, 1).value

        historia = sheet.cell(None, row, 2).value

        ano = sheet.cell(None, row, 3).value

        reitor = sheet.cell(None, row, 4).value

        localidade = sheet.cell(None, row, 5).value

        morada = sheet.cell(None, row, 6).value

        telefone = sheet.cell(None, row, 7).value

        fax = sheet.cell(None, row, 8).value

        emailUni = sheet.cell(None, row, 9).value

        emailReitor = sheet.cell(None, row, 10).value

        horarioAtendimento = sheet.cell(None, row, 11).value

        requisitosMatricula = sheet.cell(None, row, 12).value

        universidade = Universidades(nome, historia, ano, reitor, localidade, morada, telefone, fax, emailUni,
                                     emailReitor, horarioAtendimento, requisitosMatricula)

        db.session.add(universidade)



class insertEscolas:
    '''
    Inserção de Dados na Tabela Escolas apartir do ficheiro informations
    '''
    sheet = wb_data.get_sheet_by_name('Escolas_data')
    row_count = sheet.max_row

    for row in range(1, row_count + 1):
        nome = sheet.cell(None, row, 1).value

        descricao = sheet.cell(None, row, 2).value

        presidente = sheet.cell(None, row, 3).value

        vicePresidente = sheet.cell(None, row, 4).value

        conselhoCientifico = sheet.cell(None, row, 5).value

        universidade_id = sheet.cell(None, row, 6).value

        escola = Escolas(nome, descricao, presidente, vicePresidente, conselhoCientifico, universidade_id)
        db.session.add(escola)


'''
Inserção de Dados na Tabela Laboratorios apartir do ficheiro informations
'''


class insertLaboratorios:
    '''
    Inserção de Dados na Tabela Laboratorios apartir do ficheiro informations
    '''

    sheet = wb_data.get_sheet_by_name('Labs_data')
    row_count = sheet.max_row
    col_count = sheet.max_column
    for row in range(1, row_count + 1):
        nome = sheet.cell(None, row, 1).value

        quantidade = sheet.cell(None, row, 2).value

        universidade_id = sheet.cell(None, row, 3).value

        lab = Laboratorios(nome, quantidade, universidade_id)
        db.session.add(lab)

class insertCursos:
    '''
    Inserção de Dados na Tabela Cursos apartir do ficheiro informations
    '''
    sheet = wb_data.get_sheet_by_name('Cursos_data')
    row_count = sheet.max_row
    for row in range(1, row_count + 1):
        nome = sheet.cell(None, row, 1).value

        tipo = sheet.cell(None, row, 2).value

        ano = sheet.cell(None, row, 3).value

        cod_escola = sheet.cell(None, row, 4).value

        duracao_curso = sheet.cell(None, row, 5).value

        saidas_descricao = sheet.cell(None, row, 6).value

        nome_director = sheet.cell(None, row, 7).value

        email_director = sheet.cell(None, row, 8).value

        objectivos_curso = sheet.cell(None, row, 9).value

        escola_id = sheet.cell(None, row, 10).value

        curso = Cursos(nome, tipo, ano, cod_escola, duracao_curso, saidas_descricao, nome_director, email_director,
                       objectivos_curso, escola_id)

        db.session.add(curso)


class insertUnidadesCurriculares:
    '''
    Inserção de Dados na Tabela Unidades Curriculares apartir do ficheiro informations
    '''

    sheet = wb_data.get_sheet_by_name('Cadeiras_data')
    row_count = sheet.max_row
    for row in range(1, row_count + 1):
        nome = sheet.cell(None, row, 1).value

        area = sheet.cell(None, row, 2).value

        nivel = sheet.cell(None, row, 3).value

        ficha_descritiva = sheet.cell(None, row, 4).value

        disciplina_activa = sheet.cell(None, row, 5).value

        ano_curricular = sheet.cell(None, row, 6).value

        semestre = sheet.cell(None, row, 7).value

        horas_teoricas_semanal = sheet.cell(None, row, 8).value

        horas_teorica_praticas_semanal = sheet.cell(None, row, 9).value

        horas_praticas_laboratorial = sheet.cell(None, row, 10).value

        horas_trabalho_campo = sheet.cell(None, row, 11).value

        horas_seminario_semanal = sheet.cell(None, row, 12).value

        horas_estagio_semanal = sheet.cell(None, row, 13).value

        horas_orientacao_tutorial_semanal = sheet.cell(None, row, 14).value

        creditos_disciplina = sheet.cell(None, row, 15).value

        curso_id = sheet.cell(None, row, 16).value

        unidadeCurricular = unidadesCurriculares(nome, area, nivel, ficha_descritiva, disciplina_activa, ano_curricular,
                                                 semestre, horas_teoricas_semanal, horas_teorica_praticas_semanal,
                                                 horas_praticas_laboratorial, horas_trabalho_campo,
                                                 horas_seminario_semanal, horas_estagio_semanal,
                                                 horas_orientacao_tutorial_semanal, creditos_disciplina, curso_id)

        db.session.add(unidadeCurricular)



class insertProfessores:
    '''
    Inserção de Dados na Tabela Professores apartir do ficheiro informations
    '''

    sheet = wb_data.get_sheet_by_name('Professores_data')
    row_count = sheet.max_row
    for row in range(1, row_count + 1):
        nome = sheet.cell(None, row, 1).value

        email = sheet.cell(None, row, 2).value

        habilitacoes = sheet.cell(None, row, 3).value

        cadeiras_lecionadas = sheet.cell(None, row, 4).value

        responsabilidades = sheet.cell(None, row, 5).value

        actividade_profissional = sheet.cell(None, row, 6).value

        media_salario = sheet.cell(None, row, 7).value

        area_pesquisa = sheet.cell(None, row, 8).value

        licenciatura = sheet.cell(None, row, 9).value

        mestrado = sheet.cell(None, row, 10).value

        doutoramento = sheet.cell(None, row, 11).value

        professor = Professores(nome, email, habilitacoes, cadeiras_lecionadas, responsabilidades,
                                actividade_profissional, media_salario, area_pesquisa, licenciatura, mestrado,
                                doutoramento)

        db.session.add(professor)



class insertProfessoresCadeiras:
    '''
    Inserção de Dados na Tabela Professores Unidade Curricular apartir do ficheiro informations
    '''
    sheet = wb_data.get_sheet_by_name('Professores_Cadeiras_data')
    row_count = sheet.max_row
    for row in range(1, row_count + 1):
        professor = sheet.cell(None, row, 1).value

        cadeira = sheet.cell(None, row, 2).value

        professor_cadeira = unidadesCurricularesProfessores(professor, cadeira)

        db.session.add(professor_cadeira)

class insertAcessos:
    '''
    Inserção de Dados na Tabela acessoCursos apartir do ficheiro informations
    '''

    sheet = wb_data.get_sheet_by_name('Acesso_Curso_data')
    row_count = sheet.max_row
    for row in range(1, row_count + 1):
        ano = sheet.cell(None, row, 1).value

        fase = sheet.cell(None, row, 2).value

        codigo_curso = sheet.cell(None, row, 3).value

        curso = sheet.cell(None, row, 4).value

        acesso = acessoCursos(ano, fase, codigo_curso, curso)

        db.session.add(acesso)



class insertOpcoes:
    '''
    Inserção de Dados na Tabela opcoesCandidatura apartir do ficheiro informations
    '''

    sheet = wb_data.get_sheet_by_name('Opcao_Candidatura_data')
    row_count = sheet.max_row
    for row in range(1, row_count + 1):
        opcao = sheet.cell(None, row, 1).value

        candidatos = sheet.cell(None, row, 2).value

        percentagem_candidatos = sheet.cell(None, row, 3).value

        colocados = sheet.cell(None, row, 4).value

        percentagem_colocados = sheet.cell(None, row, 5).value

        acesso = sheet.cell(None, row, 6).value

        opcao = opcaoCandidatura(opcao, candidatos, percentagem_candidatos, colocados, percentagem_colocados, acesso)

        db.session.add(opcao)



class insertEtapas:
    '''
    Inserção de Dados na Tabela etapaColocacao apartir do ficheiro informations
    '''
    sheet = wb_data.get_sheet_by_name('Etapa_Colocacao_data')
    row_count = sheet.max_row
    for row in range(1, row_count + 1):
        etapa = sheet.cell(None, row, 1).value

        candidatos = sheet.cell(None, row, 2).value

        percentagem_candidatos = sheet.cell(None, row, 3).value

        colocados = sheet.cell(None, row, 4).value

        percentagem_colocados = sheet.cell(None, row, 5).value

        nota = sheet.cell(None, row, 6).value

        acesso = sheet.cell(None, row, 7).value

        etapa = etapaColocacao(etapa, candidatos, percentagem_candidatos, colocados, percentagem_colocados, nota,
                               acesso)

        db.session.add(etapa)





class insertDistritos:
    '''
    Inserção de Dados na Tabela distritoColocados apartir do ficheiro informations
    '''

    sheet = wb_data.get_sheet_by_name('Distrito_Candidatura_data')
    row_count = sheet.max_row
    for row in range(1, row_count + 1):
        distrito = sheet.cell(None, row, 1).value

        candidatos = sheet.cell(None, row, 2).value

        percentagem_candidatos = sheet.cell(None, row, 3).value

        colocados = sheet.cell(None, row, 4).value

        percentagem_colocados = sheet.cell(None, row, 5).value

        acesso = sheet.cell(None, row, 6).value

        distrito = distritoCandidatura(distrito, candidatos, percentagem_candidatos, colocados, percentagem_colocados,
                                       acesso)

        db.session.add(distrito)





class insertSexos:
    '''
    Inserção de Dados na Tabela sexoColocados apartir do ficheiro informations
    '''

    sheet = wb_data.get_sheet_by_name('Sexo_Candidatos_data')
    row_count = sheet.max_row
    for row in range(1, row_count + 1):
        sexo = sheet.cell(None, row, 1).value

        candidatos = sheet.cell(None, row, 2).value

        percentagem_candidatos = sheet.cell(None, row, 3).value

        colocados = sheet.cell(None, row, 4).value

        percentagem_colocados = sheet.cell(None, row, 5).value

        acesso = sheet.cell(None, row, 6).value

        sexo = sexoCandidatura(sexo, candidatos, percentagem_candidatos, colocados, percentagem_colocados, acesso)

        db.session.add(sexo)


class insertCursoCandidatos:
    '''
    Inserção de Dados na Tabela cursoCandidatos apartir do ficheiro informations
    '''
    sheet = wb_data.get_sheet_by_name('Cursos_Candidatos_data')
    row_count = sheet.max_row
    for row in range(1, row_count + 1):
        curso = sheet.cell(None, row, 1).value

        candidatos = sheet.cell(None, row, 2).value

        percentagem_candidatos = sheet.cell(None, row, 3).value

        colocados = sheet.cell(None, row, 4).value

        percentagem_colocados = sheet.cell(None, row, 5).value

        acesso = sheet.cell(None, row, 6).value

        curso = cursoCandidatos(curso, candidatos, percentagem_candidatos, colocados, percentagem_colocados, acesso)

        db.session.add(curso)

class insertMediaColocados:
    '''
    Inserção de Dados na Tabela Medias Colocados apartir do ficheiro informations
    '''
    sheet = wb_data.get_sheet_by_name('Medias_Colocados_data')
    row_count = sheet.max_row
    for row in range(1, row_count + 1):
        tipo_media = sheet.cell(None, row, 1).value

        media = sheet.cell(None, row, 2).value

        acesso = sheet.cell(None, row, 3).value

        media = mediaColocados(tipo_media, media, acesso)

        db.session.add(media)

class insertOpcoesExcluidas:
    '''
    Inserção de Dados na Tabela OpçoesExcluidas apartir do ficheiro informations
    '''
    sheet = wb_data.get_sheet_by_name('Opcoes_Excluidas_data')
    row_count = sheet.max_row
    for row in range(1, row_count + 1):
        tipo_exclusao = sheet.cell(None, row, 1).value

        quantidade = sheet.cell(None, row, 2).value

        acesso = sheet.cell(None, row, 3).value

        exclusao = opcoesExcluidas(tipo_exclusao, quantidade, acesso)

        db.session.add(exclusao)

class insertNotasAlunosCandidatos:
    '''
    Inserção de Dados na Tabela Medias Colocados apartir do ficheiro informations
    '''

    sheet = wb_data.get_sheet_by_name('Distribuicao_Notas_Candi_data')
    row_count = sheet.max_row
    for row in range(1, row_count + 1):
        nota = sheet.cell(None, row, 1).value

        quantidade = sheet.cell(None, row, 2).value

        acesso = sheet.cell(None, row, 3).value

        alunos = notasAlunosCandidatos(nota, quantidade, acesso)

        db.session.add(alunos)


class insertNotasAlunosColocados:
    '''
    Inserção de Dados na Tabela Medias Colocados apartir do ficheiro informations
    '''

    sheet = wb_data.get_sheet_by_name('Distribuicao_Notas_Coloc_data')
    row_count = sheet.max_row
    for row in range(1, row_count + 1):
        nota = sheet.cell(None, row, 1).value

        quantidade = sheet.cell(None, row, 2).value

        acesso = sheet.cell(None, row, 3).value

        alunos = notasAlunosColocados(nota, quantidade, acesso)

        db.session.add(alunos)

#Db session commit
db.session.commit()

#Db session.close
db.session.close()
print("Data Base 'uniMadeira.db' Created Sucessfuly!")
